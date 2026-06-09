"""
export_report.py
================
Generates a marketing manager Excel report from scored post data.

4 sheets:
  1. Summary       — key numbers, tier breakdown chart, top 3 insights
  2. What's Working — top 10 posts table, topic performance chart
  3. Quick Wins     — undervalued posts (promote) + overexposed posts (fix SEO)
  4. Content Gaps   — monthly volume chart + underperforming categories

Run after build_benchmarks.py and analyse_opportunities.py.

Usage:
    python3 reporting/export_report.py

Output:
    data/content_report.xlsx
"""

import sys
import io
import json
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent / "scoring"))
from post_value_score import compute_value_scores

DATA_PATH    = Path(__file__).parent.parent / "data" / "facility_news_posts_export.csv"
OPP_PATH     = Path(__file__).parent.parent / "data" / "opportunities.json"
OUTPUT_PATH  = Path(__file__).parent.parent / "data" / "content_report.xlsx"

# ── Colour palette ─────────────────────────────────────────────────────────
GOLD_HEX    = "B45309"
SILVER_HEX  = "475569"
BRONZE_HEX  = "92400E"
POOR_HEX    = "991B1B"
ACCENT_HEX  = "2563EB"
GREEN_HEX   = "15803D"
RED_HEX     = "991B1B"
HEADER_HEX  = "1E3A5F"
LIGHT_HEX   = "EFF6FF"
BORDER_HEX  = "E2E5EA"

TIER_COLORS = {
    "Gold":       f"#{GOLD_HEX}",
    "Silver":     f"#{SILVER_HEX}",
    "Bronze":     f"#{BRONZE_HEX}",
    "Needs Work": f"#{POOR_HEX}",
}


# ── Helpers ────────────────────────────────────────────────────────────────

def header_fill(hex_color=HEADER_HEX):
    return PatternFill("solid", fgColor=hex_color)

def light_fill():
    return PatternFill("solid", fgColor=LIGHT_HEX)

def thin_border():
    s = Side(style="thin", color=BORDER_HEX)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11):
    return Font(bold=True, color="FFFFFF", size=size)

def title_font(size=14):
    return Font(bold=True, color=HEADER_HEX, size=size)

def bold_font(size=11):
    return Font(bold=True, color="1A1D23", size=size)

def normal_font(size=10):
    return Font(color="1A1D23", size=size)

def style_header_row(ws, row, cols, hex_color=HEADER_HEX):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = header_fill(hex_color)
        cell.font   = header_font()
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data_row(ws, row, cols, shade=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if shade:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
        cell.font   = normal_font()
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def chart_to_image(fig):
    """Convert a matplotlib figure to an openpyxl Image object."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return XLImage(buf)

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def write_section_title(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = title_font()
    return row + 1


# ── Chart builders ─────────────────────────────────────────────────────────

def chart_tier_breakdown(tier_counts: dict):
    labels = ["Gold", "Silver", "Bronze", "Needs Work"]
    values = [tier_counts.get(l, 0) for l in labels]
    colors = [TIER_COLORS[l] for l in labels]

    fig, ax = plt.subplots(figsize=(5, 3.5))
    bars = ax.bar(labels, values, color=colors, width=0.5, edgecolor="white", linewidth=1.5)
    ax.set_title("Post Tier Distribution", fontsize=12, fontweight="bold", pad=12, color="#1E3A5F")
    ax.set_ylabel("Number of Posts", fontsize=9, color="#6B7280")
    ax.tick_params(colors="#6B7280", labelsize=9)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_facecolor("#FAFAFA")
    fig.patch.set_facecolor("white")
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                str(val), ha="center", va="bottom", fontsize=9, fontweight="bold", color="#1A1D23")
    fig.tight_layout()
    return chart_to_image(fig)


def chart_topic_performance(topic_avg: dict):
    topics = list(topic_avg.keys())
    scores = list(topic_avg.values())
    colors = [f"#{ACCENT_HEX}" if s >= 58 else f"#{BRONZE_HEX}" for s in scores]

    fig, ax = plt.subplots(figsize=(6, 3.8))
    bars = ax.barh(topics, scores, color=colors, edgecolor="white", linewidth=1.2)
    ax.axvline(58, color="#94A3B8", linestyle="--", linewidth=1, label="Site avg (58.3)")
    ax.set_title("Avg Value Score by Topic", fontsize=12, fontweight="bold", pad=12, color="#1E3A5F")
    ax.set_xlabel("Avg Value Score", fontsize=9, color="#6B7280")
    ax.tick_params(colors="#6B7280", labelsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_facecolor("#FAFAFA")
    fig.patch.set_facecolor("white")
    ax.legend(fontsize=8, framealpha=0.5)
    for bar, val in zip(bars, scores):
        ax.text(val + 0.5, bar.get_y() + bar.get_height() / 2,
                f"{val:.1f}", va="center", fontsize=8, color="#1A1D23")
    fig.tight_layout()
    return chart_to_image(fig)


def chart_monthly_volume(df: pd.DataFrame):
    df = df.copy()
    df["month_num"]  = pd.to_datetime(df["post_date"]).dt.month
    df["month_name"] = pd.to_datetime(df["post_date"]).dt.strftime("%b")
    counts = df.groupby(["month_num", "month_name"]).size().reset_index(name="count")
    counts = counts.sort_values("month_num")
    avg = counts["count"].mean()

    colors = [f"#{RED_HEX}" if c < avg * 0.80 else f"#{ACCENT_HEX}"
              for c in counts["count"]]

    fig, ax = plt.subplots(figsize=(7, 3.5))
    bars = ax.bar(counts["month_name"], counts["count"], color=colors,
                  edgecolor="white", linewidth=1.2)
    ax.axhline(avg, color="#94A3B8", linestyle="--", linewidth=1,
               label=f"Monthly avg ({avg:.0f})")
    ax.set_title("Posts Published per Month", fontsize=12, fontweight="bold",
                 pad=12, color="#1E3A5F")
    ax.set_ylabel("Number of Posts", fontsize=9, color="#6B7280")
    ax.tick_params(colors="#6B7280", labelsize=9)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_facecolor("#FAFAFA")
    fig.patch.set_facecolor("white")
    ax.legend(fontsize=8, framealpha=0.5)
    below = mpatches.Patch(color=f"#{RED_HEX}", label="Below average")
    above = mpatches.Patch(color=f"#{ACCENT_HEX}", label="On track")
    ax.legend(handles=[above, below], fontsize=8, framealpha=0.5)
    fig.tight_layout()
    return chart_to_image(fig)


# ── Sheet builders ─────────────────────────────────────────────────────────

def build_summary(wb: Workbook, df: pd.DataFrame, opp: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 28, "B": 18, "C": 18, "D": 18, "E": 18, "F": 5, "G": 50})

    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="Content Intelligence Report · Marketing Summary")
    cell.font = Font(bold=True, size=16, color=HEADER_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 32
    row += 1

    ws.cell(row=row, column=1, value=f"Generated: {pd.Timestamp.now().strftime('%d %B %Y')}  |  Based on {len(df)} posts (2023–2026)")
    ws.cell(row=row, column=1).font = Font(size=9, color="6B7280")
    row += 2

    # ── Key metrics ────────────────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Key Metrics")
    headers = ["Total Posts", "Avg Value Score", "Median Score", "Top 25% (≥)", "Top 10% (≥)"]
    values  = [
        len(df),
        round(df["value_score"].mean(), 1),
        round(df["value_score"].median(), 1),
        round(df["value_score"].quantile(0.75), 1),
        round(df["value_score"].quantile(0.90), 1),
    ]
    for i, (h, v) in enumerate(zip(headers, values), 1):
        hcell = ws.cell(row=row, column=i, value=h)
        hcell.fill = header_fill()
        hcell.font = header_font(10)
        hcell.border = thin_border()
        hcell.alignment = Alignment(horizontal="center")
    row += 1
    for i, v in enumerate(values, 1):
        vcell = ws.cell(row=row, column=i, value=v)
        vcell.font = Font(bold=True, size=12, color=HEADER_HEX)
        vcell.border = thin_border()
        vcell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
    row += 2

    # ── Tier breakdown chart ───────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Tier Breakdown")
    tier_counts = df["value_tier"].value_counts().to_dict()
    img = chart_tier_breakdown(tier_counts)
    img.anchor = f"A{row}"
    ws.add_image(img)
    row += 20

    # ── Top 3 insights ─────────────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Top 3 Insights")
    insights = [
        ("🏆 Best performing category",
         df.groupby("topic_category")["value_score"].mean().idxmax()),
        ("⚠️  Needs most attention",
         df.groupby("topic_category")["value_score"].mean().idxmin()),
        ("📅 Content gap month",
         "September — below-average post volume historically"),
    ]
    for label, value in insights:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = bold_font()
        lc.border = thin_border()
        lc.fill = light_fill()
        vc = ws.cell(row=row, column=2, value=value)
        ws.merge_cells(f"B{row}:E{row}")
        vc.font = normal_font()
        vc.border = thin_border()
        vc.alignment = Alignment(wrap_text=True)
        row += 1


def build_whats_working(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("What's Working")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 52, "B": 16, "C": 16, "D": 12, "E": 12, "F": 5, "G": 50})

    row = 1
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="What's Working — Top Performers")
    cell.font = title_font(14)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Top 10 posts ───────────────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Top 10 Posts by Value Score")
    import re
    df_dedup = df.copy()
    df_dedup["base_title"] = df_dedup["title"].str.replace(
        r"\s*-\s*Update\s*\d+$", "", regex=True).str.strip()
    top10 = df_dedup.sort_values("value_score", ascending=False).drop_duplicates("base_title").head(10)

    hdrs = ["Title", "Category", "Tier", "Score", "Read (min)"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill   = header_fill()
        c.font   = header_font()
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center")
    row += 1

    for idx, (_, post) in enumerate(top10.iterrows()):
        tier_color = TIER_COLORS.get(str(post["value_tier"]), "475569").lstrip("#")
        data = [
            post["base_title"],
            post["topic_category"],
            str(post["value_tier"]),
            post["value_score"],
            int(post["estimated_reading_time_minutes"]),
        ]
        for i, val in enumerate(data, 1):
            c = ws.cell(row=row, column=i, value=val)
            c.font   = normal_font()
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i == 3:
                c.fill = PatternFill("solid", fgColor=tier_color)
                c.font = Font(bold=True, color="FFFFFF", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif i == 4:
                c.font = Font(bold=True, color=HEADER_HEX, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif i == 5:
                c.alignment = Alignment(horizontal="center", vertical="center")
        if idx % 2 == 1:
            for i in range(1, 6):
                ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.row_dimensions[row].height = 28
        row += 1

    row += 1

    # ── Topic performance chart ────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Average Value Score by Topic Category")
    topic_avg = df.groupby("topic_category")["value_score"].mean().round(1).sort_values().to_dict()
    img = chart_topic_performance(topic_avg)
    img.anchor = f"A{row}"
    ws.add_image(img)


def build_quick_wins(wb: Workbook, opp: dict):
    ws = wb.create_sheet("Quick Wins")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 50, "B": 22, "C": 12, "D": 12, "E": 45})

    row = 1
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="Quick Wins — Content Opportunities")
    cell.font = title_font(14)
    ws.row_dimensions[row].height = 28
    row += 2

    for section_key, section_title, color in [
        ("undervalued", "🔍 Undervalued Posts — High SEO, Low Views (Promote These)", GREEN_HEX),
        ("overexposed",  "⚠️  Overexposed Posts — High Views, Weak SEO (Fix These)",  BRONZE_HEX),
    ]:
        row = write_section_title(ws, row, 1, section_title)
        hdrs = ["Title", "Category", "Score", "Views", "Recommended Action"]
        for i, h in enumerate(hdrs, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.fill   = PatternFill("solid", fgColor=color)
            c.font   = header_font()
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        for idx, post in enumerate(opp.get(section_key, [])):
            data = [post["title"], post["topic_category"],
                    post["value_score"], post["views"], post["action"]]
            for i, val in enumerate(data, 1):
                c = ws.cell(row=row, column=i, value=val)
                c.font   = normal_font()
                c.border = thin_border()
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if i in (3, 4):
                    c.alignment = Alignment(horizontal="center", vertical="center")
            if idx % 2 == 1:
                for i in range(1, 6):
                    ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor="F8FAFC")
            ws.row_dimensions[row].height = 30
            row += 1
        row += 2


def build_content_gaps(wb: Workbook, df: pd.DataFrame, opp: dict):
    ws = wb.create_sheet("Content Gaps")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 45, "E": 5, "F": 50})

    row = 1
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value="Content Gaps — Where to Focus Next")
    cell.font = title_font(14)
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Monthly volume chart ───────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Monthly Publishing Volume")
    img = chart_monthly_volume(df)
    img.anchor = f"A{row}"
    ws.add_image(img)
    row += 22

    # ── Underperforming topics table ───────────────────────────────────────
    row = write_section_title(ws, row, 1, "Underperforming Topic Categories")
    hdrs = ["Topic Category", "Avg Score", "vs Site Avg", "Suggested Action"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill   = header_fill(BRONZE_HEX)
        c.font   = header_font()
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    gaps = opp.get("gap_analysis", {}).get("underperforming_topics", {})
    for idx, (topic, data) in enumerate(gaps.items()):
        vals = [topic, data["avg_value_score"],
                f"{data['vs_overall']:+.1f}", data["suggestion"]]
        for i, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=val)
            c.font   = normal_font()
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i == 3:
                c.font = Font(bold=True, color=POOR_HEX, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif i == 2:
                c.alignment = Alignment(horizontal="center", vertical="center")
        if idx % 2 == 1:
            for i in range(1, 5):
                ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.row_dimensions[row].height = 28
        row += 1

    row += 2

    # ── Seasonal gaps ──────────────────────────────────────────────────────
    row = write_section_title(ws, row, 1, "Seasonal Content Gaps")
    seasonal = opp.get("gap_analysis", {}).get("seasonal_gaps", {})
    if seasonal:
        for month, data in seasonal.items():
            mc = ws.cell(row=row, column=1, value=month)
            mc.font = bold_font()
            mc.border = thin_border()
            mc.fill = PatternFill("solid", fgColor="FEF2F2")
            sc = ws.cell(row=row, column=2,
                         value=f"{data['posts']} posts (avg {data['avg']:.0f})")
            sc.font   = normal_font()
            sc.border = thin_border()
            sc.alignment = Alignment(horizontal="center")
            ac = ws.cell(row=row, column=3, value=data["suggestion"])
            ws.merge_cells(f"C{row}:D{row}")
            ac.font   = normal_font()
            ac.border = thin_border()
            ac.alignment = Alignment(wrap_text=True)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No significant seasonal gaps detected.").font = normal_font()


# ── Main ───────────────────────────────────────────────────────────────────

def export():
    print(f"Loading {DATA_PATH} …")
    raw = pd.read_csv(DATA_PATH)
    df  = compute_value_scores(raw)
    print(f"Scored {len(df)} posts.")

    if not OPP_PATH.exists():
        print("⚠️  opportunities.json not found — run analyse_opportunities.py first.")
        return
    opp = json.loads(OPP_PATH.read_text())

    wb = Workbook()
    build_summary(wb, df, opp)
    build_whats_working(wb, df)
    build_quick_wins(wb, opp)
    build_content_gaps(wb, df, opp)

    wb.save(OUTPUT_PATH)
    print(f"\n✅  content_report.xlsx written → {OUTPUT_PATH}")
    print(f"    Sheets: Summary, What's Working, Quick Wins, Content Gaps")


if __name__ == "__main__":
    export()
